import openpyxl


class GetPasswordFromExcel:
    def __init__(self, file):
        self.file = file
        self.password_list = []
        self.resolve_data = self.__load_data_sheet_from_file()
        self.data = self.__load_data_from_file()

    def __load_data_sheet_from_file(self):
        try:
            data = openpyxl.load_workbook(self.file)
            data_sheet = data.worksheets[0]
        except FileExistsError as e:
            print(e)
        else:
            return data_sheet

    def __load_data_from_file(self):
        return openpyxl.load_workbook(self.file)

    def get_first_blank_row_index(self):
        i = 1
        while True:
            if self.resolve_data.cell(row=i, column=1).value is None:
                break
            i += 1
        return i

    def get_passwords(self):
        i = 1
        while True:
            if self.resolve_data.cell(row=i, column=1).value is None:
                break
            self.password_list.append(self.resolve_data.cell(row=i, column=1).value)
            i += 1
        return self.password_list


class SavePasswordToExcel:
    file_name = 'dane.xlsx'

    def __init__(self, password_list):
        self.password_list = password_list
        self.index = GetPasswordFromExcel(self.file_name)

    def get_first_blank_row(self):
        return self.index.get_first_blank_row_index()

    def save_password_to_excel(self):
        data_sheet = self.index.resolve_data
        data = self.index.data
        for password in self.password_list:
            data_sheet.cell(row=self.get_first_blank_row(), column=1).value = password
            data.save(self.file_name)


if __name__ == "__main__":
    passwords_list = GetPasswordFromExcel('dane.xlsx')
    print(passwords_list.get_passwords())
    first_blank = SavePasswordToExcel(["a"])
    print(first_blank.get_first_blank_row())
    print(first_blank.save_password_to_excel())
    openpyxl.Workbook
